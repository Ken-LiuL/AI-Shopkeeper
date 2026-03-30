"""Excel export API routes for products, orders, and inventory."""

from __future__ import annotations

import io
import logging
from datetime import datetime

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse

from src.db import postgres as pg

router = APIRouter(prefix="/api/export", tags=["export"])
logger = logging.getLogger(__name__)

_EXCEL_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _make_excel_response(wb, filename: str) -> StreamingResponse:
    """Serialize workbook to StreamingResponse."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=_EXCEL_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get("/products")
async def export_products() -> StreamingResponse:
    """Export all products as Excel.

    Columns: product_id, name, brand, category, retail_price,
             cost_price, status, monthly_sales, stock
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    pool = pg.get_pool()
    rows = await pool.fetch(
        """
        SELECT
            product_id,
            name,
            brand,
            category,
            retail_price,
            cost_price,
            status,
            monthly_sales,
            stock
        FROM products
        ORDER BY name
        """
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品列表"

    headers = [
        "商品ID", "商品名称", "品牌", "类目",
        "零售价", "成本价", "状态", "月销量", "库存",
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in rows:
        ws.append([
            r["product_id"],
            r["name"],
            r["brand"] or "",
            r["category"] or "",
            float(r["retail_price"]) if r["retail_price"] is not None else "",
            float(r["cost_price"]) if r["cost_price"] is not None else "",
            r["status"] or "",
            r["monthly_sales"] or 0,
            r["stock"] or 0,
        ])

    # Auto-fit column widths (approximate)
    col_widths = [15, 30, 15, 15, 12, 12, 10, 10, 10]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    today = datetime.now().strftime("%Y%m%d")
    return _make_excel_response(wb, f"products_{today}.xlsx")


@router.get("/orders")
async def export_orders(
    start_date: str | None = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: str | None = Query(None, description="结束日期 YYYY-MM-DD"),
) -> StreamingResponse:
    """Export orders as Excel.

    Columns: order_id, order_time, status, total_amount, item_count
    Supports optional date range filtering via start_date / end_date.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    pool = pg.get_pool()

    conditions: list[str] = []
    params: list[str] = []

    if start_date:
        params.append(start_date)
        conditions.append(f"order_time >= ${len(params)}::date")
    if end_date:
        params.append(end_date)
        conditions.append(f"order_time < (${len(params)}::date + INTERVAL '1 day')")

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    # Fetch orders with item count via LEFT JOIN
    rows = await pool.fetch(
        f"""
        SELECT
            o.order_id,
            o.order_time,
            o.status,
            o.total_amount,
            COUNT(oi.id)::int AS item_count
        FROM orders o
        LEFT JOIN order_items oi ON oi.order_id = o.order_id
        {where}
        GROUP BY o.order_id, o.order_time, o.status, o.total_amount
        ORDER BY o.order_time DESC NULLS LAST
        """,
        *params,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单列表"

    headers = ["订单ID", "下单时间", "状态", "订单金额", "商品数量"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in rows:
        order_time = r["order_time"]
        if order_time is not None:
            order_time_str = order_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            order_time_str = ""
        ws.append([
            r["order_id"],
            order_time_str,
            r["status"] or "",
            float(r["total_amount"]) if r["total_amount"] is not None else "",
            r["item_count"] or 0,
        ])

    col_widths = [20, 20, 15, 14, 10]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    today = datetime.now().strftime("%Y%m%d")
    suffix = ""
    if start_date or end_date:
        suffix = f"_{start_date or 'start'}_{end_date or 'end'}"
    return _make_excel_response(wb, f"orders{suffix}_{today}.xlsx")


@router.get("/inventory")
async def export_inventory() -> StreamingResponse:
    """Export inventory as Excel.

    Columns: sku, name, stock, available_stock, safety_stock
    Pulls from qnh_inventory (latest snapshot) joined with product names,
    with fallback to the products table.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    pool = pg.get_pool()

    # Try qnh_inventory first (most recent snapshot per sku)
    rows = await pool.fetch(
        """
        SELECT DISTINCT ON (COALESCE(NULLIF(qi.sku_id, ''), qi.spu_id))
            COALESCE(NULLIF(qi.sku_id, ''), qi.spu_id)          AS sku,
            COALESCE(p.name, qi.product_name, '')               AS name,
            qi.current_stock                                     AS stock,
            qi.available_stock,
            NULL::int                                            AS safety_stock
        FROM qnh_inventory qi
        LEFT JOIN products p
               ON p.product_id = COALESCE(NULLIF(qi.sku_id, ''), qi.spu_id)
        ORDER BY COALESCE(NULLIF(qi.sku_id, ''), qi.spu_id), qi.snapshot_time DESC
        """
    )

    # Fallback: use products table if qnh_inventory is empty
    if not rows:
        rows = await pool.fetch(
            """
            SELECT
                product_id       AS sku,
                name,
                stock,
                NULL::int        AS available_stock,
                NULL::int        AS safety_stock
            FROM products
            ORDER BY name
            """
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存列表"

    headers = ["SKU", "商品名称", "库存数量", "可用库存", "安全库存"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in rows:
        ws.append([
            r["sku"] or "",
            r["name"] or "",
            r["stock"] if r["stock"] is not None else "",
            r["available_stock"] if r["available_stock"] is not None else "",
            r["safety_stock"] if r["safety_stock"] is not None else "",
        ])

    col_widths = [20, 30, 12, 12, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    today = datetime.now().strftime("%Y%m%d")
    return _make_excel_response(wb, f"inventory_{today}.xlsx")
