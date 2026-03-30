"""Manual Excel import service for products, orders, and inventory."""

from __future__ import annotations

import contextlib
import json
import logging
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

import openpyxl
import xlrd

logger = logging.getLogger(__name__)


def _build_issue_key(issue_type: str, row: dict[str, Any]) -> str:
    normalized = {key: row[key] for key in sorted(row.keys())}
    return f"{issue_type}:{json.dumps(normalized, ensure_ascii=False, sort_keys=True, separators=(',', ':'))}"


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"-", "--", "无", "None"}:
        return ""
    return text


def _clean_header(value: Any) -> str:
    return re.sub(r"\s+", "", _clean_text(value))


def _parse_decimal(value: Any) -> Decimal | None:
    raw = _clean_text(value)
    if not raw:
        return None
    raw = raw.replace(",", "").replace("¥", "")
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _parse_int(value: Any) -> int | None:
    raw = _clean_text(value)
    if not raw:
        return None
    raw = raw.replace(",", "")
    try:
        return int(float(raw))
    except ValueError:
        return None


def _parse_datetime(value: Any) -> datetime | None:
    raw = _clean_text(value)
    if not raw:
        return None
    candidates = (
        "%Y-%m-%d %H:%M:%S",
        "%Y.%m.%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y.%m.%d",
    )
    for fmt in candidates:
        try:
            return datetime.strptime(raw, fmt).replace(tzinfo=UTC)
        except ValueError:
            continue
    return None


def _normalize_status(raw: str, kind: str) -> str:
    text = _clean_text(raw)
    if kind == "product":
        if text in {"售卖", "在售", "上架"}:
            return "active"
        if text in {"停售", "下架", "停用"}:
            return "inactive"
        return "active" if not text else text.lower()
    if kind == "order":
        mapping = {
            "已完成": "completed",
            "配送中": "processing",
            "拣货完成": "processing",
            "骑手已接单": "processing",
            "骑手已送达": "processing",
            "等待分配骑手": "pending",
            "已取消": "cancelled",
            "退款成功": "refunded",
        }
        return mapping.get(text, text or "pending")
    return text


def _quality_score(total_rows: int, weighted_issues: int) -> float:
    if total_rows <= 0:
        return 100.0
    penalty = min(95.0, (weighted_issues / max(total_rows, 1)) * 100)
    return round(max(5.0, 100.0 - penalty), 2)


@dataclass
class QualityIssue:
    severity: str
    code: str
    message: str
    count: int = 1
    samples: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "code": self.code,
            "message": self.message,
            "count": self.count,
            "samples": self.samples[:5],
        }


@dataclass
class ImportPreview:
    import_type: str
    filename: str
    detected_sheets: list[str]
    total_rows: int
    normalized_preview: dict[str, list[dict[str, Any]]]
    quality_report: dict[str, Any]


@dataclass
class ImportResult:
    run_id: str
    import_type: str
    filename: str
    detected_sheets: list[str]
    total_rows: int
    imported_rows: int
    skipped_rows: int
    quality_report: dict[str, Any]
    import_summary: dict[str, Any]


class ManualImportService:
    PRODUCT_HEADERS = {"SPU编码", "SKU编码", "商品名称", "总部零售价"}
    ORDER_LIST_HEADERS = {"订单号", "门店编码", "下单时间", "订单总金额"}
    ORDER_DETAIL_HEADERS = {"订单号", "商品名称", "商品SKU码", "商品销售数量"}
    INVENTORY_HEADERS = {"SKU", "商品名称", "总库存", "可用库存"}

    def __init__(self, pool: Any):
        self.pool = pool

    def preview(self, filename: str, content: bytes, import_type: str | None = None) -> ImportPreview:
        workbook = self._load_workbook(filename, content)
        detected_type = import_type or self._detect_import_type(workbook)
        parser = {
            "products": self._parse_products,
            "orders": self._parse_orders,
            "inventory": self._parse_inventory,
        }.get(detected_type)
        if parser is None:
            raise ValueError(f"Unsupported import type: {detected_type}")
        parsed = parser(workbook)
        return ImportPreview(
            import_type=detected_type,
            filename=filename,
            detected_sheets=parsed["detected_sheets"],
            total_rows=parsed["total_rows"],
            normalized_preview=parsed["preview"],
            quality_report=parsed["quality_report"],
        )

    async def import_file(
        self,
        filename: str,
        content: bytes,
        import_type: str | None = None,
    ) -> ImportResult:
        workbook = self._load_workbook(filename, content)
        detected_type = import_type or self._detect_import_type(workbook)
        parser = {
            "products": self._parse_products,
            "orders": self._parse_orders,
            "inventory": self._parse_inventory,
        }.get(detected_type)
        if parser is None:
            raise ValueError(f"Unsupported import type: {detected_type}")
        parsed = parser(workbook)
        run_id = f"imp_{datetime.now(UTC).strftime('%Y%m%d%H%M%S%f')}"
        imported_rows = 0
        skipped_rows = 0
        summary: dict[str, Any]

        if detected_type == "products":
            imported_rows, skipped_rows, summary = await self._commit_products(parsed["records"])
        elif detected_type == "orders":
            imported_rows, skipped_rows, summary = await self._commit_orders(parsed["records"])
        else:
            imported_rows, skipped_rows, summary = await self._commit_inventory(parsed["records"])

        review_snapshot = await self.get_review(limit=20)
        summary = {
            **summary,
            "review_snapshot": {
                "summary": review_snapshot.get("summary", {}),
                "open_summary": review_snapshot.get("open_summary", {}),
            },
        }

        await self._record_run(
            run_id=run_id,
            import_type=detected_type,
            filename=filename,
            detected_sheets=parsed["detected_sheets"],
            total_rows=parsed["total_rows"],
            imported_rows=imported_rows,
            skipped_rows=skipped_rows,
            quality_report=parsed["quality_report"],
            import_summary=summary,
        )

        # ── 导入完成后自动跑派生 ETL ──────────────────────────────
        if imported_rows > 0:
            await self._run_post_import_etl(detected_type)

        return ImportResult(
            run_id=run_id,
            import_type=detected_type,
            filename=filename,
            detected_sheets=parsed["detected_sheets"],
            total_rows=parsed["total_rows"],
            imported_rows=imported_rows,
            skipped_rows=skipped_rows,
            quality_report=parsed["quality_report"],
            import_summary=summary,
        )

    async def list_runs(self, limit: int = 20) -> list[dict[str, Any]]:
        rows = await self.pool.fetch(
            """
            SELECT run_id, import_type, filename, status, total_rows, imported_rows,
                   skipped_rows, quality_score, quality_report, import_summary,
                   created_at, updated_at
            FROM manual_import_runs
            ORDER BY created_at DESC
            LIMIT $1
            """,
            limit,
        )
        return [dict(row) for row in rows]

    async def get_overview(self) -> dict[str, Any]:
        latest = await self.pool.fetchrow(
            """
            SELECT import_type, filename, quality_score, created_at
            FROM manual_import_runs
            ORDER BY created_at DESC
            LIMIT 1
            """
        )
        rows = await self.pool.fetch(
            """
            SELECT import_type,
                   COUNT(*)::int AS run_count,
                   COALESCE(SUM(imported_rows), 0)::int AS imported_rows
            FROM manual_import_runs
            GROUP BY import_type
            """
        )
        return {
            "latest_run": dict(latest) if latest else None,
            "by_type": [dict(row) for row in rows],
        }

    async def get_run_comparison(self, import_type: str | None = None) -> dict[str, Any]:
        params: list[Any] = []
        conditions = ["status = 'completed'"]

        if import_type:
            conditions.append("import_type = $1")
            params.append(import_type)
        else:
            latest_type = await self.pool.fetchval(
                """
                SELECT import_type
                FROM manual_import_runs
                WHERE status = 'completed'
                ORDER BY created_at DESC NULLS LAST
                LIMIT 1
                """
            )
            if latest_type:
                import_type = str(latest_type)
                conditions.append("import_type = $1")
                params.append(import_type)

        rows = await self.pool.fetch(
            f"""
            SELECT run_id, import_type, filename, status, total_rows, imported_rows,
                   skipped_rows, quality_score, quality_report, import_summary, created_at, updated_at
            FROM manual_import_runs
            WHERE {' AND '.join(conditions)}
            ORDER BY created_at DESC NULLS LAST
            LIMIT 2
            """,
            *params,
        )

        latest = dict(rows[0]) if rows else None
        previous = dict(rows[1]) if len(rows) > 1 else None

        if not latest:
            return {
                "import_type": import_type,
                "latest_run": None,
                "previous_run": None,
                "delta": {},
                "review_delta": {
                    "available": False,
                    "new_issues": [],
                    "resolved_issues": [],
                    "worsened_issues": [],
                    "improved_issues": [],
                },
            }

        def _review_snapshot(run: dict[str, Any] | None) -> dict[str, Any] | None:
            if not run:
                return None
            import_summary = run.get("import_summary") or {}
            if isinstance(import_summary, str):
                with contextlib.suppress(Exception):
                    import_summary = json.loads(import_summary)
            snapshot = import_summary.get("review_snapshot") if isinstance(import_summary, dict) else None
            return snapshot if isinstance(snapshot, dict) else None

        latest_snapshot = _review_snapshot(latest)
        previous_snapshot = _review_snapshot(previous)
        latest_open = (latest_snapshot or {}).get("open_summary") or {}
        previous_open = (previous_snapshot or {}).get("open_summary") or {}

        issue_labels = {
            "stockout_but_selling": "断货但仍有销量",
            "catalog_gaps": "商品主档缺口",
            "products_missing_price": "商品缺售价",
            "order_amount_mismatch": "订单金额异常",
            "inventory_missing_cost": "库存缺成本价",
        }
        new_issues: list[dict[str, Any]] = []
        resolved_issues: list[dict[str, Any]] = []
        worsened_issues: list[dict[str, Any]] = []
        improved_issues: list[dict[str, Any]] = []

        if latest_open and previous_open:
            for key in sorted(set(latest_open) | set(previous_open)):
                current = int(latest_open.get(key) or 0)
                previous_value = int(previous_open.get(key) or 0)
                delta = current - previous_value
                payload = {
                    "key": key,
                    "title": issue_labels.get(key, key),
                    "current": current,
                    "previous": previous_value,
                    "delta": delta,
                }
                if previous_value == 0 and current > 0:
                    new_issues.append(payload)
                elif current == 0 and previous_value > 0:
                    resolved_issues.append(payload)
                elif delta > 0:
                    worsened_issues.append(payload)
                elif delta < 0:
                    improved_issues.append(payload)

        return {
            "import_type": import_type or latest.get("import_type"),
            "latest_run": latest,
            "previous_run": previous,
            "delta": {
                "imported_rows": int(latest.get("imported_rows") or 0) - int((previous or {}).get("imported_rows") or 0),
                "total_rows": int(latest.get("total_rows") or 0) - int((previous or {}).get("total_rows") or 0),
                "quality_score": round(float(latest.get("quality_score") or 0) - float((previous or {}).get("quality_score") or 0), 2),
                "open_issues": int(sum(latest_open.values()) if latest_open else 0) - int(sum(previous_open.values()) if previous_open else 0),
            },
            "review_delta": {
                "available": bool(latest_open and previous_open),
                "new_issues": new_issues,
                "resolved_issues": resolved_issues,
                "worsened_issues": worsened_issues,
                "improved_issues": improved_issues,
            },
        }

    async def get_review(self, limit: int = 20) -> dict[str, Any]:
        tables = {
            name: await self._table_exists(name)
            for name in (
                "products",
                "orders",
                "order_items",
                "qnh_inventory",
                "qnh_products",
                "product_knowledge",
                "manual_import_runs",
                "qnh_dataset_records",
                "issue_actions",
            )
        }
        summary = {
            "products": await self.pool.fetchval("SELECT COUNT(*) FROM products") if tables["products"] else 0,
            "orders": await self.pool.fetchval("SELECT COUNT(*) FROM orders") if tables["orders"] else 0,
            "order_items": await self.pool.fetchval("SELECT COUNT(*) FROM order_items") if tables["order_items"] else 0,
            "inventory_rows": await self.pool.fetchval("SELECT COUNT(*) FROM qnh_inventory") if tables["qnh_inventory"] else 0,
            "product_knowledge": await self.pool.fetchval("SELECT COUNT(*) FROM product_knowledge") if tables["product_knowledge"] else 0,
            "import_runs": await self.pool.fetchval("SELECT COUNT(*) FROM manual_import_runs") if tables["manual_import_runs"] else 0,
            "products_with_sales": await self.pool.fetchval(
                "SELECT COUNT(*) FROM products WHERE COALESCE(monthly_sales, 0) > 0"
            ) if tables["products"] else 0,
            "products_with_stock": await self.pool.fetchval(
                "SELECT COUNT(*) FROM products WHERE COALESCE(stock, 0) > 0"
            ) if tables["products"] else 0,
            "products_missing_price": await self.pool.fetchval(
                "SELECT COUNT(*) FROM products WHERE COALESCE(retail_price, 0) = 0"
            ) if tables["products"] else 0,
            "stockout_but_selling": await self.pool.fetchval(
                """
                SELECT COUNT(*)
                FROM products
                WHERE COALESCE(stock, 0) = 0
                  AND COALESCE(monthly_sales, 0) > 0
                """
            ) if tables["products"] else 0,
            "inventory_missing_cost": await self.pool.fetchval(
                "SELECT COUNT(*) FROM qnh_inventory WHERE cost_price IS NULL"
            ) if tables["qnh_inventory"] else 0,
            "catalog_gaps": await self.pool.fetchval(
                """
                SELECT COUNT(*)
                FROM products p
                LEFT JOIN product_knowledge pk
                  ON pk.sku_id = COALESCE(NULLIF(p.sku_id, ''), p.product_id)
                 AND pk.spu_id = COALESCE(NULLIF(p.spu_id, ''), COALESCE(NULLIF(p.sku_id, ''), p.product_id))
                WHERE pk.id IS NULL
                """
            ) if tables["products"] and tables["product_knowledge"] else 0,
            "order_amount_mismatch": await self.pool.fetchval(
                """
                SELECT COUNT(*)
                FROM (
                    SELECT o.order_id
                    FROM orders o
                    JOIN order_items oi ON oi.order_id = o.order_id
                    GROUP BY o.order_id, o.customer_paid
                    HAVING ABS(COALESCE(SUM(oi.quantity * oi.unit_price), 0) - COALESCE(o.customer_paid, 0)) > 5
                ) t
                """
            ) if tables["orders"] and tables["order_items"] else 0,
        }

        stockout_rows = await self.pool.fetch(
            """
            SELECT product_id, name, stock, monthly_sales, retail_price
            FROM products
            WHERE COALESCE(stock, 0) = 0
              AND COALESCE(monthly_sales, 0) > 0
            ORDER BY monthly_sales DESC, name ASC
            LIMIT $1
            """,
            limit,
        ) if tables["products"] else []
        missing_price_rows = await self.pool.fetch(
            """
            SELECT product_id, name, category, stock, monthly_sales
            FROM products
            WHERE COALESCE(retail_price, 0) = 0
            ORDER BY monthly_sales DESC, name ASC
            LIMIT $1
            """,
            limit,
        ) if tables["products"] else []
        catalog_gap_rows = await self.pool.fetch(
            """
            SELECT p.product_id, p.sku_id, p.name, p.stock, p.monthly_sales, p.source
            FROM products p
            LEFT JOIN product_knowledge pk
              ON pk.sku_id = COALESCE(NULLIF(p.sku_id, ''), p.product_id)
             AND pk.spu_id = COALESCE(NULLIF(p.spu_id, ''), COALESCE(NULLIF(p.sku_id, ''), p.product_id))
            WHERE pk.id IS NULL
            ORDER BY p.monthly_sales DESC, p.stock DESC, p.name ASC
            LIMIT $1
            """,
            limit,
        ) if tables["products"] and tables["product_knowledge"] else []
        mismatch_rows = await self.pool.fetch(
            """
            SELECT o.order_id,
                   o.status,
                   o.order_time,
                   COALESCE(o.customer_paid, 0) AS customer_paid,
                   COALESCE(SUM(oi.quantity * oi.unit_price), 0)::numeric(10,2) AS line_total,
                   ABS(COALESCE(SUM(oi.quantity * oi.unit_price), 0) - COALESCE(o.customer_paid, 0))::numeric(10,2) AS diff
            FROM orders o
            JOIN order_items oi ON oi.order_id = o.order_id
            GROUP BY o.order_id, o.status, o.order_time, o.customer_paid
            HAVING ABS(COALESCE(SUM(oi.quantity * oi.unit_price), 0) - COALESCE(o.customer_paid, 0)) > 5
            ORDER BY diff DESC, o.order_time DESC NULLS LAST
            LIMIT $1
            """,
            limit,
        ) if tables["orders"] and tables["order_items"] else []
        inventory_cost_rows = await self.pool.fetch(
            """
            SELECT sku_id, product_name, store_name, stock
            FROM qnh_inventory
            WHERE cost_price IS NULL
            ORDER BY stock DESC, product_name ASC
            LIMIT $1
            """,
            limit,
        ) if tables["qnh_inventory"] else []

        open_summary = dict(summary)
        closed_issue_keys: dict[str, set[str]] = defaultdict(set)
        if tables["issue_actions"]:
            try:
                action_rows = await self.pool.fetch(
                    """
                    SELECT issue_type, issue_key
                    FROM issue_actions
                    WHERE issue_type = ANY($1::text[])
                      AND status IN ('resolved', 'ignored')
                    """,
                    [
                        "product_missing_price",
                        "product_catalog_gap",
                        "order_amount_mismatch",
                        "stockout_but_selling",
                        "inventory_missing_cost",
                    ],
                )
                for row in action_rows:
                    closed_issue_keys[str(row["issue_type"])].add(str(row["issue_key"]))
            except Exception:
                logger.debug("Failed to load issue action statuses", exc_info=True)

        if closed_issue_keys.get("product_missing_price"):
            all_missing_price_rows = await self.pool.fetch(
                """
                SELECT product_id, name, category, stock, monthly_sales
                FROM products
                WHERE COALESCE(retail_price, 0) = 0
                ORDER BY monthly_sales DESC, name ASC
                """
            ) if tables["products"] else []
            open_missing = [
                dict(row)
                for row in all_missing_price_rows
                if _build_issue_key("product_missing_price", dict(row))
                not in closed_issue_keys["product_missing_price"]
            ]
            open_summary["products_missing_price"] = len(open_missing)
            missing_price_rows = open_missing[:limit]

        if closed_issue_keys.get("product_catalog_gap"):
            all_catalog_gap_rows = await self.pool.fetch(
                """
                SELECT p.product_id, p.sku_id, p.name, p.stock, p.monthly_sales, p.source
                FROM products p
                LEFT JOIN product_knowledge pk
                  ON pk.sku_id = COALESCE(NULLIF(p.sku_id, ''), p.product_id)
                 AND pk.spu_id = COALESCE(NULLIF(p.spu_id, ''), COALESCE(NULLIF(p.sku_id, ''), p.product_id))
                WHERE pk.id IS NULL
                ORDER BY p.monthly_sales DESC, p.stock DESC, p.name ASC
                """
            ) if tables["products"] and tables["product_knowledge"] else []
            open_catalog = [
                dict(row)
                for row in all_catalog_gap_rows
                if _build_issue_key("product_catalog_gap", dict(row))
                not in closed_issue_keys["product_catalog_gap"]
            ]
            open_summary["catalog_gaps"] = len(open_catalog)
            catalog_gap_rows = open_catalog[:limit]

        if closed_issue_keys.get("order_amount_mismatch"):
            all_mismatch_rows = await self.pool.fetch(
                """
                SELECT o.order_id,
                       o.status,
                       o.order_time,
                       COALESCE(o.customer_paid, 0) AS customer_paid,
                       COALESCE(SUM(oi.quantity * oi.unit_price), 0)::numeric(10,2) AS line_total,
                       ABS(COALESCE(SUM(oi.quantity * oi.unit_price), 0) - COALESCE(o.customer_paid, 0))::numeric(10,2) AS diff
                FROM orders o
                JOIN order_items oi ON oi.order_id = o.order_id
                GROUP BY o.order_id, o.status, o.order_time, o.customer_paid
                HAVING ABS(COALESCE(SUM(oi.quantity * oi.unit_price), 0) - COALESCE(o.customer_paid, 0)) > 5
                ORDER BY diff DESC, o.order_time DESC NULLS LAST
                """
            ) if tables["orders"] and tables["order_items"] else []
            open_mismatch = [
                dict(row)
                for row in all_mismatch_rows
                if _build_issue_key("order_amount_mismatch", dict(row))
                not in closed_issue_keys["order_amount_mismatch"]
            ]
            open_summary["order_amount_mismatch"] = len(open_mismatch)
            mismatch_rows = open_mismatch[:limit]

        if closed_issue_keys.get("stockout_but_selling"):
            all_stockout_rows = await self.pool.fetch(
                """
                SELECT product_id, name, stock, monthly_sales, retail_price
                FROM products
                WHERE COALESCE(stock, 0) = 0
                  AND COALESCE(monthly_sales, 0) > 0
                ORDER BY monthly_sales DESC, name ASC
                """
            ) if tables["products"] else []
            open_stockout = [
                dict(row)
                for row in all_stockout_rows
                if _build_issue_key("stockout_but_selling", dict(row))
                not in closed_issue_keys["stockout_but_selling"]
            ]
            open_summary["stockout_but_selling"] = len(open_stockout)
            stockout_rows = open_stockout[:limit]

        if closed_issue_keys.get("inventory_missing_cost"):
            all_inventory_cost_rows = await self.pool.fetch(
                """
                SELECT sku_id, product_name, store_name, stock
                FROM qnh_inventory
                WHERE cost_price IS NULL
                ORDER BY stock DESC, product_name ASC
                """
            ) if tables["qnh_inventory"] else []
            open_inventory_cost = [
                dict(row)
                for row in all_inventory_cost_rows
                if _build_issue_key("inventory_missing_cost", dict(row))
                not in closed_issue_keys["inventory_missing_cost"]
            ]
            open_summary["inventory_missing_cost"] = len(open_inventory_cost)
            inventory_cost_rows = open_inventory_cost[:limit]

        issues = [
            {
                "key": "stockout_but_selling",
                "title": "有销量但库存为 0",
                "severity": "critical",
                "count": int(open_summary["stockout_but_selling"] or 0),
                "description": "这些商品在近 30 天仍有销量，但当前库存已经归零。",
                "recommended_action": "优先补货或核对库存导出逻辑，避免继续缺货损失。",
            },
            {
                "key": "catalog_gaps",
                "title": "商品主档缺口",
                "severity": "warning",
                "count": int(open_summary["catalog_gaps"] or 0),
                "description": "订单或库存里出现了商品，但商品主档/商品知识里没有完整记录。",
                "recommended_action": "补齐商品规格明细表，避免客服知识和商品分析失真。",
            },
            {
                "key": "products_missing_price",
                "title": "商品缺少零售价",
                "severity": "warning",
                "count": int(open_summary["products_missing_price"] or 0),
                "description": "缺失零售价会削弱定价、毛利和客单价分析。",
                "recommended_action": "补齐总部零售价或渠道售价，再做价格相关分析。",
            },
            {
                "key": "order_amount_mismatch",
                "title": "订单金额与明细金额差异较大",
                "severity": "warning",
                "count": int(open_summary["order_amount_mismatch"] or 0),
                "description": "这通常意味着订单级优惠、配送费或明细拆分规则需要单独处理。",
                "recommended_action": "利润分析先排除这类订单，后续再补贴优惠拆分口径。",
            },
            {
                "key": "inventory_missing_cost",
                "title": "库存缺少成本价",
                "severity": "info",
                "count": int(open_summary["inventory_missing_cost"] or 0),
                "description": "缺失成本价不会阻塞库存功能，但会影响库存价值和毛利分析。",
                "recommended_action": "优先补热销和高库存商品的成本价。",
            },
        ]

        return {
            "summary": summary,
            "open_summary": open_summary,
            "issues": issues,
            "tables": {
                "stockout_but_selling": [dict(row) for row in stockout_rows],
                "missing_price": [dict(row) for row in missing_price_rows],
                "catalog_gaps": [dict(row) for row in catalog_gap_rows],
                "order_amount_mismatch": [dict(row) for row in mismatch_rows],
                "inventory_missing_cost": [dict(row) for row in inventory_cost_rows],
            },
        }

    async def _table_exists(self, table_name: str) -> bool:
        exists = await self.pool.fetchval(
            """
            SELECT EXISTS (
                SELECT 1
                FROM information_schema.tables
                WHERE table_schema = 'public' AND table_name = $1
            )
            """,
            table_name,
        )
        return bool(exists)

    def _load_workbook(self, filename: str, content: bytes) -> dict[str, list[dict[str, Any]]]:
        suffix = filename.lower().rsplit(".", 1)[-1]
        if suffix == "xlsx":
            workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
            result: dict[str, list[dict[str, Any]]] = {}
            for ws in workbook.worksheets:
                ws.reset_dimensions()
                rows = ws.iter_rows(values_only=True)
                header_row = next(rows, ())
                headers = [_clean_header(cell) for cell in header_row]
                data_rows: list[dict[str, Any]] = []
                for values in rows:
                    if not any(_clean_text(value) for value in values):
                        continue
                    record = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
                    data_rows.append(record)
                result[ws.title] = data_rows
            return result
        if suffix == "xls":
            book = xlrd.open_workbook(file_contents=content)
            result = {}
            for index in range(book.nsheets):
                sheet = book.sheet_by_index(index)
                headers = [_clean_header(value) for value in sheet.row_values(0)]
                data_rows = []
                for row_idx in range(1, sheet.nrows):
                    values = sheet.row_values(row_idx)
                    if not any(_clean_text(value) for value in values):
                        continue
                    record = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
                    data_rows.append(record)
                result[sheet.name] = data_rows
            return result
        raise ValueError("Only .xlsx and .xls files are supported")

    def _detect_import_type(self, workbook: dict[str, list[dict[str, Any]]]) -> str:
        header_sets = {name: set(rows[0].keys()) if rows else set() for name, rows in workbook.items()}
        for headers in header_sets.values():
            if self.PRODUCT_HEADERS.issubset(headers):
                return "products"
        if any(self.ORDER_LIST_HEADERS.issubset(headers) for headers in header_sets.values()):
            return "orders"
        if any(self.INVENTORY_HEADERS.issubset(headers) for headers in header_sets.values()):
            return "inventory"
        raise ValueError("Unable to detect import type from workbook headers")

    def _parse_products(self, workbook: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
        sheet_name = next((name for name, rows in workbook.items() if rows and self.PRODUCT_HEADERS.issubset(rows[0].keys())), None)
        if sheet_name is None:
            raise ValueError("Product sheet not found")
        product_rows = workbook[sheet_name]
        combo_name = next((name for name in workbook if "组合" in name), None)
        combo_rows = workbook.get(combo_name or "", [])
        combo_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in combo_rows:
            sku_id = _clean_text(row.get("skuId"))
            child_id = _clean_text(row.get("子商品ID"))
            factor = _parse_decimal(row.get("组合系数"))
            if sku_id and child_id:
                combo_map[sku_id].append(
                    {
                        "child_product_id": child_id,
                        "factor": float(factor) if factor is not None else 0.0,
                    }
                )

        records = []
        issues: list[QualityIssue] = []
        missing_price = 0
        missing_category = 0
        missing_barcode = 0
        duplicate_keys: set[str] = set()
        seen_keys: set[str] = set()

        for row in product_rows:
            spu_id = _clean_text(row.get("SPU编码"))
            sku_id = _clean_text(row.get("SKU编码"))
            name = _clean_text(row.get("商品名称"))
            barcode = _clean_text(row.get("商品条码"))
            category = _clean_text(row.get("店内分类")) or _clean_text(row.get("美团渠道一级类目"))
            retail_price = _parse_decimal(row.get("总部零售价"))
            cost_price = _parse_decimal(row.get("成本单价"))
            description_parts = [
                _clean_text(row.get("商品描述")),
                _clean_text(row.get("商品卖点")),
                _clean_text(row.get("商品名称补充语")),
            ]
            description = "；".join(part for part in description_parts if part)
            product_id = sku_id or spu_id
            if not product_id or not name:
                continue
            if not retail_price:
                missing_price += 1
            if not category:
                missing_category += 1
            if not barcode:
                missing_barcode += 1
            uniq_key = f"{spu_id}:{sku_id}"
            if uniq_key in seen_keys:
                duplicate_keys.add(uniq_key)
            seen_keys.add(uniq_key)
            records.append(
                {
                    "product_id": product_id,
                    "spu_id": spu_id or product_id,
                    "sku_id": sku_id or product_id,
                    "name": name,
                    "barcode": barcode,
                    "upc_code": barcode,
                    "category": category,
                    "category_leaf": _clean_text(row.get("店内分类末级分类")),
                    "brand": _clean_text(row.get("品牌名称")) or _clean_text(row.get("京东品牌名称")),
                    "spec": _clean_text(row.get("规格名称")),
                    "description": description,
                    "retail_price": retail_price,
                    "cost_price": cost_price,
                    "weight": _parse_decimal(row.get("重量")),
                    "weight_unit": _clean_text(row.get("重量单位")),
                    "status": _normalize_status(_clean_text(row.get("规格售卖状态")) or _clean_text(row.get("商品售卖状态")), "product"),
                    "product_type": _clean_text(row.get("商品类型")),
                    "is_standard": _clean_text(row.get("是否标品")) == "标品",
                    "is_batch_managed": _clean_text(row.get("批次管理")) == "是",
                    "shelf_life_days": _parse_int(row.get("保质期（天）")),
                    "source_created_at": _parse_datetime(row.get("商品创建时间")),
                    "cover_store_count": _parse_int(row.get("覆盖门店数")) or 0,
                    "selling_points": _clean_text(row.get("商品卖点")),
                    "extra": {
                        "channel_prices": {
                            "eleme": float(_parse_decimal(row.get("饿了么渠道售价")) or 0),
                            "meituan": float(_parse_decimal(row.get("美团外卖渠道售价")) or 0),
                            "jddj": float(_parse_decimal(row.get("京东秒送渠道售价")) or 0),
                        },
                        "tags": {
                            "spu": _clean_text(row.get("SPU商品标签")),
                            "sku": _clean_text(row.get("SKU商品标签")),
                        },
                        "combo_children": combo_map.get(sku_id or product_id, []),
                    },
                }
            )

        if missing_price:
            issues.append(QualityIssue("warning", "missing_price", "商品缺少零售价", missing_price))
        if missing_category:
            issues.append(QualityIssue("warning", "missing_category", "商品缺少分类", missing_category))
        if missing_barcode:
            issues.append(QualityIssue("info", "missing_barcode", "商品缺少条码", missing_barcode))
        if duplicate_keys:
            issues.append(
                QualityIssue(
                    "critical",
                    "duplicate_sku",
                    "商品表中存在重复 SPU/SKU 组合",
                    len(duplicate_keys),
                    samples=sorted(duplicate_keys)[:5],
                )
            )

        quality_report = self._build_quality_report(
            total_rows=len(records),
            issues=issues,
            stats={
                "records": len(records),
                "missing_price": missing_price,
                "missing_category": missing_category,
                "missing_barcode": missing_barcode,
                "combo_relations": sum(len(v) for v in combo_map.values()),
            },
            suggestions=[
                "建议优先补齐缺失价格和分类的商品，否则定价、选品和低库存判断会变弱。",
                "商品描述和卖点会直接进入自动补全知识文本，用于客服和商品检索。",
            ],
        )
        return {
            "records": {"products": records},
            "preview": {"products": records[:5]},
            "detected_sheets": [sheet_name] + ([combo_name] if combo_name else []),
            "quality_report": quality_report,
            "total_rows": len(records),
        }

    def _parse_orders(self, workbook: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
        list_name = next((name for name, rows in workbook.items() if rows and self.ORDER_LIST_HEADERS.issubset(rows[0].keys())), None)
        detail_name = next((name for name, rows in workbook.items() if rows and self.ORDER_DETAIL_HEADERS.issubset(rows[0].keys())), None)
        if list_name is None:
            raise ValueError("Order list sheet not found")
        order_rows = workbook[list_name]
        detail_rows = workbook.get(detail_name or "", [])

        items_by_order: dict[str, list[dict[str, Any]]] = defaultdict(list)
        missing_item_sku = 0
        unmatched_detail_orders: set[str] = set()
        detail_total_mismatch = 0

        for row in detail_rows:
            order_id = _clean_text(row.get("订单号"))
            sku_id = _clean_text(row.get("商品SKU码"))
            quantity = _parse_int(row.get("商品销售数量")) or 0
            unit_price = _parse_decimal(row.get("商品售价")) or Decimal("0")
            total_price = _parse_decimal(row.get("商品总售价")) or Decimal("0")
            if not sku_id:
                missing_item_sku += 1
            item = {
                "product_id": sku_id or _clean_text(row.get("商品条码")) or _clean_text(row.get("商品名称")),
                "sku_id": sku_id,
                "barcode": _clean_text(row.get("商品条码")),
                "name": _clean_text(row.get("商品名称")),
                "spec": _clean_text(row.get("商品规格")),
                "category": _clean_text(row.get("店内分类")),
                "quantity": quantity,
                "unit_price": unit_price,
                "line_total": total_price,
                "refunded": _clean_text(row.get("是否退款商品")) == "是",
                "refund_qty": _parse_int(row.get("退款数量")) or 0,
                "order_time": _parse_datetime(row.get("下单时间")),
                "completed_time": _parse_datetime(row.get("订单完成时间")),
            }
            if order_id:
                items_by_order[order_id].append(item)

        records = []
        seen_orders: set[str] = set()
        duplicate_orders: set[str] = set()
        missing_paid = 0

        for row in order_rows:
            order_id = _clean_text(row.get("订单号"))
            if not order_id:
                continue
            if order_id in seen_orders:
                duplicate_orders.add(order_id)
            seen_orders.add(order_id)
            order_time = _parse_datetime(row.get("下单时间"))
            paid_amount = _parse_decimal(row.get("支付金额"))
            total_amount = _parse_decimal(row.get("订单总金额"))
            if paid_amount is None:
                missing_paid += 1
            store_id = _clean_text(row.get("门店编码"))
            status = _normalize_status(_clean_text(row.get("订单状态")), "order")
            item_rows = items_by_order.get(order_id, [])
            if detail_name and not item_rows:
                unmatched_detail_orders.add(order_id)
            line_sum = sum((item["line_total"] for item in item_rows), Decimal("0"))
            if paid_amount is not None and item_rows and abs(line_sum - paid_amount) > Decimal("5"):
                detail_total_mismatch += 1
            merchant_discount = sum(
                (
                    _parse_decimal(row.get("整单商家优惠")) or Decimal("0"),
                    _parse_decimal(row.get("单品商家优惠")) or Decimal("0"),
                    _parse_decimal(row.get("配送费商家优惠")) or Decimal("0"),
                ),
                Decimal("0"),
            )
            delivery_fee = _parse_decimal(row.get("履约服务费")) or Decimal("0")
            order = {
                "order_id": order_id,
                "platform": self._normalize_platform(_clean_text(row.get("渠道"))),
                "store_id": store_id,
                "store_name": _clean_text(row.get("门店名称")),
                "customer_name": _clean_text(row.get("收件人")),
                "customer_phone_suffix": self._extract_phone_suffix(
                    _clean_text(row.get("收件人真实电话")) or _clean_text(row.get("收件人虚拟号"))
                ),
                "total_amount": total_amount or Decimal("0"),
                "customer_paid": paid_amount or Decimal("0"),
                "status": status,
                "order_time": order_time,
                "order_date": order_time.date() if order_time else None,
                "delivery_address_type": _clean_text(row.get("订单时效类型")),
                "commission": _parse_decimal(row.get("商品佣金")) or Decimal("0"),
                "delivery_fee": delivery_fee,
                "merchant_discount": merchant_discount,
                "day_seq": _parse_int(row.get("订单流水号")),
                "store_city": _clean_text(row.get("门店所属城市")),
                "delivery_status": _clean_text(row.get("配送状态")),
                "delivery_method": _clean_text(row.get("配送方式")),
                "items": item_rows,
                "extra": {
                    "merchant_income": float(_parse_decimal(row.get("商家预计收入")) or 0),
                    "package_fee_customer": float(_parse_decimal(row.get("顾客支付包装费")) or 0),
                    "package_fee_merchant": float(_parse_decimal(row.get("商家包装费收入")) or 0),
                    "order_flags": _clean_text(row.get("订单标识")),
                    "refund_reason": _clean_text(row.get("用户申请退款原因")),
                    "cancel_reason": _clean_text(row.get("订单取消原因")),
                    "remark": _clean_text(row.get("备注")),
                },
            }
            records.append(order)

        issues: list[QualityIssue] = []
        if duplicate_orders:
            issues.append(
                QualityIssue(
                    "critical",
                    "duplicate_order_id",
                    "订单列表中存在重复订单号",
                    len(duplicate_orders),
                    samples=sorted(duplicate_orders)[:5],
                )
            )
        if missing_paid:
            issues.append(QualityIssue("warning", "missing_paid_amount", "订单缺少支付金额", missing_paid))
        if missing_item_sku:
            issues.append(QualityIssue("warning", "missing_item_sku", "订单明细缺少 SKU", missing_item_sku))
        if unmatched_detail_orders:
            issues.append(
                QualityIssue(
                    "warning",
                    "orders_without_detail",
                    "订单列表中部分订单没有匹配到订单明细",
                    len(unmatched_detail_orders),
                    samples=sorted(unmatched_detail_orders)[:5],
                )
            )
        if detail_total_mismatch:
            issues.append(
                QualityIssue(
                    "info",
                    "order_item_total_mismatch",
                    "部分订单明细金额与支付金额差异较大",
                    detail_total_mismatch,
                )
            )

        quality_report = self._build_quality_report(
            total_rows=len(records) + len(detail_rows),
            issues=issues,
            stats={
                "orders": len(records),
                "order_items": len(detail_rows),
                "missing_paid_amount": missing_paid,
                "missing_item_sku": missing_item_sku,
                "orders_without_detail": len(unmatched_detail_orders),
                "detail_total_mismatch": detail_total_mismatch,
            },
            suggestions=[
                "订单导入会同步回填 products.monthly_sales，并生成订单商品行供补货、套餐、日报使用。",
                "若订单状态和配送状态长期缺失，订单履约分析和客服订单上下文会受影响。",
            ],
        )
        return {
            "records": {
                "orders": records,
                "order_items": detail_rows,
            },
            "preview": {
                "orders": records[:5],
                "order_items": list(items_by_order.values())[:3],
            },
            "detected_sheets": [name for name in (list_name, detail_name) if name],
            "quality_report": quality_report,
            "total_rows": len(records) + len(detail_rows),
        }

    def _parse_inventory(self, workbook: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
        summary_name = next((name for name, rows in workbook.items() if rows and self.INVENTORY_HEADERS.issubset(rows[0].keys())), None)
        if summary_name is None:
            raise ValueError("Inventory summary sheet not found")
        detail_name = next((name for name in workbook if "库位库存明细" in name), None)
        detail_rows = workbook.get(detail_name or "", [])
        location_index: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in detail_rows:
            sku_id = _clean_text(row.get("SKU"))
            if sku_id:
                location_index[sku_id].append(
                    {
                        "location": _clean_text(row.get("库位")),
                        "stock": _parse_int(row.get("库位库存")) or 0,
                    }
                )

        records = []
        negative_stock = 0
        missing_cost = 0
        inconsistent_available = 0
        for row in workbook[summary_name]:
            sku_id = _clean_text(row.get("SKU"))
            name = _clean_text(row.get("商品名称"))
            if not sku_id or not name:
                continue
            total_stock = _parse_int(row.get("总库存")) or 0
            available_stock = _parse_int(row.get("可用库存")) or 0
            locked_stock = _parse_int(row.get("锁定库存")) or 0
            if total_stock < 0 or available_stock < 0 or locked_stock < 0:
                negative_stock += 1
            if available_stock > total_stock:
                inconsistent_available += 1
            if _parse_decimal(row.get("成本单价")) is None:
                missing_cost += 1
            records.append(
                {
                    "store_id": _clean_text(row.get("门店id/仓库编码")),
                    "store_name": _clean_text(row.get("门店/仓库名称")),
                    "product_name": name,
                    "sku_id": sku_id,
                    "barcode": _clean_text(row.get("UPC")),
                    "external_code": _clean_text(row.get("外部货品编码")),
                    "spec": _clean_text(row.get("规格")),
                    "category": _clean_text(row.get("品类")),
                    "warehouse": _clean_text(row.get("库位")),
                    "current_stock": total_stock,
                    "stock": total_stock,
                    "available_stock": available_stock,
                    "locked_stock": locked_stock,
                    "in_transit_stock": _parse_int(row.get("在途库存")) or 0,
                    "cost_price": _parse_decimal(row.get("成本单价")),
                    "stock_value": _parse_decimal(row.get("成本总价")),
                    "weight": _parse_decimal(row.get("重量")),
                    "weight_unit": "g",
                    "is_batch_managed": _clean_text(row.get("是否为批次商品")) == "是",
                    "is_bundle": _clean_text(row.get("组合品标识")) == "是",
                    "location_details": location_index.get(sku_id, []),
                    "extra": {
                        "channel_stock": {
                            "meituan": _parse_int(row.get("美团外卖(渠道库存)")) or 0,
                            "eleme": _parse_int(row.get("饿了么(渠道库存)")) or 0,
                            "jddj": _parse_int(row.get("京东到家(渠道库存)")) or 0,
                        },
                        "sync_by_sellout": _clean_text(row.get("按售罄同步渠道")),
                        "unlimited_online_stock": _clean_text(row.get("线上无限库存")) == "是",
                        "dimensions_cm": {
                            "length": float(_parse_decimal(row.get("长（cm）")) or 0),
                            "width": float(_parse_decimal(row.get("宽（cm）")) or 0),
                            "height": float(_parse_decimal(row.get("高（cm）")) or 0),
                        },
                        "max_spec_weight_kg": float(_parse_decimal(row.get("最大规格重量(kg)")) or 0),
                    },
                }
            )

        issues: list[QualityIssue] = []
        if negative_stock:
            issues.append(QualityIssue("critical", "negative_stock", "库存出现负数", negative_stock))
        if inconsistent_available:
            issues.append(
                QualityIssue(
                    "warning",
                    "available_gt_total",
                    "可用库存大于总库存，建议核对库存逻辑",
                    inconsistent_available,
                )
            )
        if missing_cost:
            issues.append(QualityIssue("info", "missing_cost_price", "库存缺少成本价", missing_cost))
        quality_report = self._build_quality_report(
            total_rows=len(records),
            issues=issues,
            stats={
                "inventory_rows": len(records),
                "location_rows": len(detail_rows),
                "negative_stock": negative_stock,
                "available_gt_total": inconsistent_available,
                "missing_cost_price": missing_cost,
                "zero_stock_items": sum(1 for item in records if item["stock"] == 0),
            },
            suggestions=[
                "库存导入会优先回填 qnh_inventory 和 products.stock，使库存页、补货和告警立即可用。",
                "库位库存明细会进入 extra，用于后续人工盘点和库位差异排查。",
            ],
        )
        return {
            "records": {"inventory": records},
            "preview": {"inventory": records[:5]},
            "detected_sheets": [name for name in (summary_name, detail_name) if name],
            "quality_report": quality_report,
            "total_rows": len(records),
        }

    async def _commit_products(self, payload: dict[str, list[dict[str, Any]]]) -> tuple[int, int, dict[str, Any]]:
        products = payload["products"]
        imported = 0
        skipped = 0
        async with self.pool.acquire() as conn:
            async with conn.transaction():
                for item in products:
                    if not item["product_id"]:
                        skipped += 1
                        continue
                    await conn.execute(
                        """
                        INSERT INTO products (
                            product_id, spu_id, sku_id, name, barcode, upc_code, category, brand,
                            description, cost_price, retail_price, stock, monthly_sales, status,
                            store_id, image_url, source, extra, created_at, updated_at
                        ) VALUES (
                            $1, $2, $3, $4, NULLIF($5, ''), NULLIF($6, ''), NULLIF($7, ''),
                            NULLIF($8, ''), NULLIF($9, ''), $10, $11, 0, 0, $12, NULL, NULL,
                            'manual_import', $13::jsonb, NOW(), NOW()
                        )
                        ON CONFLICT (product_id) DO UPDATE SET
                            spu_id = EXCLUDED.spu_id,
                            sku_id = EXCLUDED.sku_id,
                            name = EXCLUDED.name,
                            barcode = COALESCE(EXCLUDED.barcode, products.barcode),
                            upc_code = COALESCE(EXCLUDED.upc_code, products.upc_code),
                            category = COALESCE(NULLIF(EXCLUDED.category, ''), products.category),
                            brand = COALESCE(NULLIF(EXCLUDED.brand, ''), products.brand),
                            description = COALESCE(NULLIF(EXCLUDED.description, ''), products.description),
                            cost_price = COALESCE(EXCLUDED.cost_price, products.cost_price),
                            retail_price = COALESCE(EXCLUDED.retail_price, products.retail_price),
                            status = EXCLUDED.status,
                            source = 'manual_import',
                            extra = COALESCE(products.extra, '{}'::jsonb) || EXCLUDED.extra,
                            updated_at = NOW()
                        """,
                        item["product_id"],
                        item["spu_id"],
                        item["sku_id"],
                        item["name"],
                        item["barcode"],
                        item["upc_code"],
                        item["category"],
                        item["brand"],
                        item["description"],
                        item["cost_price"],
                        item["retail_price"],
                        item["status"],
                        json.dumps(item["extra"], ensure_ascii=False),
                    )
                    await conn.execute(
                        """
                        INSERT INTO qnh_products (
                            spu_id, sku_id, name, barcode, category, brand, spec,
                            cost_price, retail_price, status, unit, extra, synced_at,
                            stock, monthly_sales, updated_at
                        ) VALUES (
                            $1, $2, $3, NULLIF($4, ''), NULLIF($5, ''), NULLIF($6, ''), NULLIF($7, ''),
                            $8, $9, $10, NULLIF($11, ''), $12::jsonb, NOW(),
                            0, 0, NOW()
                        )
                        ON CONFLICT (spu_id, sku_id) DO UPDATE SET
                            name = EXCLUDED.name,
                            barcode = COALESCE(EXCLUDED.barcode, qnh_products.barcode),
                            category = COALESCE(NULLIF(EXCLUDED.category, ''), qnh_products.category),
                            brand = COALESCE(NULLIF(EXCLUDED.brand, ''), qnh_products.brand),
                            spec = COALESCE(NULLIF(EXCLUDED.spec, ''), qnh_products.spec),
                            cost_price = COALESCE(EXCLUDED.cost_price, qnh_products.cost_price),
                            retail_price = COALESCE(EXCLUDED.retail_price, qnh_products.retail_price),
                            status = EXCLUDED.status,
                            unit = COALESCE(NULLIF(EXCLUDED.unit, ''), qnh_products.unit),
                            extra = COALESCE(qnh_products.extra, '{}'::jsonb) || EXCLUDED.extra,
                            updated_at = NOW(),
                            synced_at = NOW()
                        """,
                        item["spu_id"],
                        item["sku_id"],
                        item["name"],
                        item["barcode"],
                        item["category"],
                        item["brand"],
                        item["spec"],
                        item["cost_price"],
                        item["retail_price"],
                        "在售" if item["status"] == "active" else "停售",
                        item["weight_unit"],
                        json.dumps(
                            {
                                "description": item["description"],
                                "selling_points": item["selling_points"],
                                "is_standard": item["is_standard"],
                                "is_batch_managed": item["is_batch_managed"],
                                "shelf_life_days": item["shelf_life_days"],
                                **item["extra"],
                            },
                            ensure_ascii=False,
                        ),
                    )
                    await self._upsert_product_knowledge(conn, item)
                    imported += 1
        return imported, skipped, {"products_upserted": imported}

    async def _commit_orders(self, payload: dict[str, list[dict[str, Any]]]) -> tuple[int, int, dict[str, Any]]:
        orders = payload["orders"]
        imported_orders = 0
        imported_items = 0
        skipped = 0
        async with self.pool.acquire() as conn:
            async with conn.transaction():
                order_ids = [item["order_id"] for item in orders if item.get("order_id")]
                if order_ids:
                    await conn.execute("DELETE FROM order_items WHERE order_id = ANY($1::varchar[])", order_ids)
                for order in orders:
                    if not order["order_id"]:
                        skipped += 1
                        continue
                    await conn.execute(
                        """
                        INSERT INTO orders (
                            order_id, platform, customer_phone_suffix, total_amount, status,
                            order_time, delivery_address_type, created_at, store_id, customer_name,
                            customer_paid, order_date, commission, delivery_fee, merchant_discount,
                            day_seq, items, source
                        ) VALUES (
                            $1, $2, NULLIF($3, ''), $4, $5, $6, NULLIF($7, ''), NOW(), NULLIF($8, ''),
                            NULLIF($9, ''), $10, $11, $12, $13, $14, $15, $16::jsonb, 'manual_import'
                        )
                        ON CONFLICT (order_id) DO UPDATE SET
                            platform = EXCLUDED.platform,
                            customer_phone_suffix = EXCLUDED.customer_phone_suffix,
                            total_amount = EXCLUDED.total_amount,
                            status = EXCLUDED.status,
                            order_time = EXCLUDED.order_time,
                            delivery_address_type = EXCLUDED.delivery_address_type,
                            store_id = EXCLUDED.store_id,
                            customer_name = EXCLUDED.customer_name,
                            customer_paid = EXCLUDED.customer_paid,
                            order_date = EXCLUDED.order_date,
                            commission = EXCLUDED.commission,
                            delivery_fee = EXCLUDED.delivery_fee,
                            merchant_discount = EXCLUDED.merchant_discount,
                            day_seq = EXCLUDED.day_seq,
                            items = EXCLUDED.items,
                            source = 'manual_import'
                        """,
                        order["order_id"],
                        order["platform"],
                        order["customer_phone_suffix"],
                        order["total_amount"],
                        order["status"],
                        order["order_time"],
                        order["delivery_address_type"],
                        order["store_id"],
                        order["customer_name"],
                        order["customer_paid"],
                        order["order_date"],
                        order["commission"],
                        order["delivery_fee"],
                        order["merchant_discount"],
                        order["day_seq"],
                        json.dumps(
                            {"products": order["items"], "extra": order["extra"]},
                            ensure_ascii=False,
                            default=str,
                        ),
                    )
                    await conn.execute(
                        """
                        INSERT INTO qnh_orders (
                            order_id, channel, store_name, total_amount, paid_amount, status,
                            order_time, delivery_fee, packaging_fee, customer_phone_suffix, items,
                            extra, synced_at
                        ) VALUES (
                            $1, $2, NULLIF($3, ''), $4, $5, $6, $7, $8, 0,
                            NULLIF($9, ''), $10::jsonb, $11::jsonb, NOW()
                        )
                        ON CONFLICT (order_id) DO UPDATE SET
                            channel = EXCLUDED.channel,
                            store_name = EXCLUDED.store_name,
                            total_amount = EXCLUDED.total_amount,
                            paid_amount = EXCLUDED.paid_amount,
                            status = EXCLUDED.status,
                            order_time = EXCLUDED.order_time,
                            delivery_fee = EXCLUDED.delivery_fee,
                            customer_phone_suffix = EXCLUDED.customer_phone_suffix,
                            items = EXCLUDED.items,
                            extra = EXCLUDED.extra,
                            synced_at = NOW()
                        """,
                        order["order_id"],
                        order["platform"],
                        order["store_name"],
                        order["total_amount"],
                        order["customer_paid"],
                        order["status"],
                        order["order_time"],
                        order["delivery_fee"],
                        order["customer_phone_suffix"],
                        json.dumps(order["items"], ensure_ascii=False, default=str),
                        json.dumps(order["extra"], ensure_ascii=False, default=str),
                    )
                    for item in order["items"]:
                        product_id = item["product_id"]
                        if not product_id:
                            skipped += 1
                            continue
                        await self._ensure_product_stub(conn, item)
                        await conn.execute(
                            """
                            INSERT INTO order_items (order_id, product_id, quantity, unit_price, created_at)
                            VALUES ($1, $2, $3, $4, NOW())
                            ON CONFLICT (order_id, product_id) DO UPDATE SET
                                quantity = EXCLUDED.quantity,
                                unit_price = EXCLUDED.unit_price
                            """,
                            order["order_id"],
                            product_id,
                            item["quantity"] or 0,
                            item["unit_price"] or Decimal("0"),
                        )
                        imported_items += 1
                    imported_orders += 1
                await self._refresh_sales_metrics(conn)
                await self._rebuild_hotsale_dataset(conn)
        return imported_orders + imported_items, skipped, {
            "orders_upserted": imported_orders,
            "order_items_inserted": imported_items,
        }

    async def _commit_inventory(self, payload: dict[str, list[dict[str, Any]]]) -> tuple[int, int, dict[str, Any]]:
        items = payload["inventory"]
        imported = 0
        skipped = 0
        aggregated_items = self._aggregate_inventory_records(items)
        async with self.pool.acquire() as conn:
            async with conn.transaction():
                store_ids = sorted({item["store_id"] for item in items if item["store_id"]})
                if store_ids:
                    await conn.execute(
                        "DELETE FROM qnh_inventory WHERE store_id = ANY($1::varchar[])",
                        store_ids,
                    )
                for item in items:
                    if not item["sku_id"]:
                        skipped += 1
                        continue
                    await conn.execute(
                        """
                        INSERT INTO qnh_inventory (
                            store_id, store_name, sku_id, barcode, product_name, current_stock,
                            available_stock, locked_stock, cost_price, stock_value, warehouse,
                            snapshot_time, extra, synced_at, stock, updated_at
                        ) VALUES (
                            NULLIF($1, ''), NULLIF($2, ''), $3, NULLIF($4, ''), $5, $6, $7, $8,
                            $9, $10, NULLIF($11, ''), NOW(), $12::jsonb, NOW(), $13, NOW()
                        )
                        """,
                        item["store_id"],
                        item["store_name"],
                        item["sku_id"],
                        item["barcode"],
                        item["product_name"],
                        item["current_stock"],
                        item["available_stock"],
                        item["locked_stock"],
                        item["cost_price"],
                        item["stock_value"],
                        item["warehouse"],
                        json.dumps(
                            {
                                "category": item["category"],
                                "spec": item["spec"],
                                "external_code": item["external_code"],
                                "location_details": item["location_details"],
                                **item["extra"],
                            },
                            ensure_ascii=False,
                        ),
                        item["stock"],
                    )
                    imported += 1

                for item in aggregated_items:
                    resolved_spu_id = await conn.fetchval(
                        """
                        SELECT COALESCE(
                            (
                                SELECT NULLIF(spu_id, '')
                                FROM products
                                WHERE sku_id = $1 OR product_id = $1
                                LIMIT 1
                            ),
                            (
                                SELECT NULLIF(spu_id, '')
                                FROM qnh_products
                                WHERE sku_id = $1
                                LIMIT 1
                            ),
                            $1
                        )
                        """,
                        item["sku_id"],
                    )
                    await conn.execute(
                        """
                        INSERT INTO products (
                            product_id, sku_id, name, barcode, upc_code, category, cost_price,
                            stock, status, store_id, source, extra, created_at, updated_at
                        ) VALUES (
                            $1, $2, $3, NULLIF($4, ''), NULLIF($5, ''), NULLIF($6, ''), $7,
                            $8, 'active', NULLIF($9, ''), 'manual_import', $10::jsonb, NOW(), NOW()
                        )
                        ON CONFLICT (product_id) DO UPDATE SET
                            name = EXCLUDED.name,
                            barcode = COALESCE(EXCLUDED.barcode, products.barcode),
                            upc_code = COALESCE(EXCLUDED.upc_code, products.upc_code),
                            category = COALESCE(NULLIF(EXCLUDED.category, ''), products.category),
                            cost_price = COALESCE(EXCLUDED.cost_price, products.cost_price),
                            stock = EXCLUDED.stock,
                            store_id = COALESCE(EXCLUDED.store_id, products.store_id),
                            source = 'manual_import',
                            extra = COALESCE(products.extra, '{}'::jsonb) || EXCLUDED.extra,
                            updated_at = NOW()
                        """,
                        item["sku_id"],
                        item["sku_id"],
                        item["product_name"],
                        item["barcode"],
                        item["barcode"],
                        item["category"],
                        item["cost_price"],
                        item["stock"],
                        item["store_id"],
                        json.dumps(
                            {
                                "spec": item["spec"],
                                "warehouse_names": item["warehouses"],
                                "store_names": item["store_names"],
                                "store_ids": item["store_ids"],
                            },
                            ensure_ascii=False,
                        ),
                    )
                    await conn.execute(
                        """
                        INSERT INTO qnh_products (
                            spu_id, sku_id, name, barcode, category, spec, cost_price,
                            status, extra, synced_at, stock, stock_num, monthly_sales, updated_at
                        ) VALUES (
                            $1, $2, $3, NULLIF($4, ''), NULLIF($5, ''), NULLIF($6, ''), $7,
                            '在售', $8::jsonb, NOW(), $9, $9, 0, NOW()
                        )
                        ON CONFLICT (spu_id, sku_id) DO UPDATE SET
                            name = COALESCE(NULLIF(EXCLUDED.name, ''), qnh_products.name),
                            barcode = COALESCE(EXCLUDED.barcode, qnh_products.barcode),
                            category = COALESCE(NULLIF(EXCLUDED.category, ''), qnh_products.category),
                            spec = COALESCE(NULLIF(EXCLUDED.spec, ''), qnh_products.spec),
                            cost_price = COALESCE(EXCLUDED.cost_price, qnh_products.cost_price),
                            stock = EXCLUDED.stock,
                            stock_num = EXCLUDED.stock_num,
                            extra = COALESCE(qnh_products.extra, '{}'::jsonb) || EXCLUDED.extra,
                            synced_at = NOW(),
                            updated_at = NOW()
                        """,
                        resolved_spu_id,
                        item["sku_id"],
                        item["product_name"],
                        item["barcode"],
                        item["category"],
                        item["spec"],
                        item["cost_price"],
                        json.dumps(
                            {
                                "inventory_store_ids": item["store_ids"],
                                "inventory_store_names": item["store_names"],
                                "inventory_warehouses": item["warehouses"],
                            },
                            ensure_ascii=False,
                        ),
                        item["stock"],
                    )
        return imported, skipped, {"inventory_rows_upserted": imported}

    async def _ensure_product_stub(self, conn: Any, item: dict[str, Any]) -> None:
        if not item["product_id"]:
            return
        await conn.execute(
            """
            INSERT INTO products (
                product_id, sku_id, name, barcode, upc_code, category, description,
                retail_price, stock, status, source, created_at, updated_at
            ) VALUES (
                $1, NULLIF($2, ''), $3, NULLIF($4, ''), NULLIF($5, ''), NULLIF($6, ''), NULLIF($7, ''),
                $8, 0, 'active', 'manual_import', NOW(), NOW()
            )
            ON CONFLICT (product_id) DO UPDATE SET
                name = COALESCE(NULLIF(EXCLUDED.name, ''), products.name),
                barcode = COALESCE(EXCLUDED.barcode, products.barcode),
                upc_code = COALESCE(EXCLUDED.upc_code, products.upc_code),
                category = COALESCE(NULLIF(EXCLUDED.category, ''), products.category),
                description = COALESCE(NULLIF(EXCLUDED.description, ''), products.description),
                retail_price = COALESCE(EXCLUDED.retail_price, products.retail_price),
                source = 'manual_import',
                updated_at = NOW()
            """,
            item["product_id"],
            item["sku_id"],
            item["name"],
            item["barcode"],
            item["barcode"],
            item["category"],
            item["spec"],
            item["unit_price"],
        )

    async def _refresh_sales_metrics(self, conn: Any) -> None:
        await conn.execute(
            """
            UPDATE products
            SET monthly_sales = 0,
                updated_at = NOW()
            WHERE COALESCE(monthly_sales, 0) <> 0
            """
        )
        await conn.execute(
            """
            UPDATE products p
            SET monthly_sales = COALESCE(s.qty, 0),
                updated_at = NOW()
            FROM (
                SELECT oi.product_id, COALESCE(SUM(oi.quantity), 0)::int AS qty
                FROM order_items oi
                JOIN orders o ON o.order_id = oi.order_id
                WHERE o.order_time >= NOW() - INTERVAL '30 days'
                GROUP BY oi.product_id
            ) s
            WHERE p.product_id = s.product_id
            """
        )
        await conn.execute(
            """
            UPDATE qnh_products
            SET monthly_sales = 0,
                updated_at = NOW(),
                synced_at = NOW()
            WHERE COALESCE(monthly_sales, 0) <> 0
            """
        )
        await conn.execute(
            """
            UPDATE qnh_products qp
            SET monthly_sales = s.qty,
                updated_at = NOW(),
                synced_at = NOW()
            FROM (
                SELECT oi.product_id, COALESCE(SUM(oi.quantity), 0)::int AS qty
                FROM order_items oi
                JOIN orders o ON o.order_id = oi.order_id
                WHERE o.order_time >= NOW() - INTERVAL '30 days'
                GROUP BY oi.product_id
            ) s
            WHERE qp.sku_id = s.product_id
               OR qp.spu_id = s.product_id
            """
        )
        await conn.execute(
            """
            INSERT INTO product_sales (product_id, sale_date, quantity, revenue)
            SELECT oi.product_id,
                   COALESCE(o.order_date, DATE(o.order_time), CURRENT_DATE) AS sale_date,
                   COALESCE(SUM(oi.quantity), 0)::int AS quantity,
                   COALESCE(SUM(oi.quantity * oi.unit_price), 0)::numeric(10,2) AS revenue
            FROM order_items oi
            JOIN orders o ON o.order_id = oi.order_id
            GROUP BY oi.product_id, COALESCE(o.order_date, DATE(o.order_time), CURRENT_DATE)
            ON CONFLICT (product_id, sale_date) DO UPDATE SET
                quantity = EXCLUDED.quantity,
                revenue = EXCLUDED.revenue
            """
        )

    async def _rebuild_hotsale_dataset(self, conn: Any) -> None:
        await conn.execute(
            """
            DELETE FROM qnh_dataset_records
            WHERE dataset = 'hotsale_goods'
            """
        )
        rows = await conn.fetch(
            """
            SELECT p.product_id, p.name,
                   COALESCE(SUM(oi.quantity), 0)::int AS qty,
                   COALESCE(SUM(oi.quantity * oi.unit_price), 0)::numeric(10,2) AS revenue,
                   COALESCE(MAX(p.retail_price), 0) AS price
            FROM products p
            JOIN order_items oi ON oi.product_id = p.product_id
            JOIN orders o ON o.order_id = oi.order_id
            WHERE o.order_time >= NOW() - INTERVAL '30 days'
            GROUP BY p.product_id, p.name
            ORDER BY revenue DESC, qty DESC
            LIMIT 200
            """
        )
        for rank, row in enumerate(rows, start=1):
            payload = {
                "product_id": row["product_id"],
                "product_name": row["name"],
                "prod_sale_amt": str(row["revenue"] or 0),
                "prod_sale_num_gmv": str(row["qty"] or 0),
                "prod_actual_pay_amt": str(row["revenue"] or 0),
                "rank": str(rank),
                "price": str(row["price"] or 0),
            }
            await conn.execute(
                """
                INSERT INTO qnh_dataset_records (
                    date, metric_name, metric_value, source, synced_at, dataset, payload, created_at
                ) VALUES (
                    CURRENT_DATE, $1, 0, 'manual_import', NOW(), 'hotsale_goods', $2::jsonb, NOW()
                )
                """,
                f"hotsale_{row['product_id']}",
                json.dumps(payload, ensure_ascii=False),
            )

    async def _upsert_product_knowledge(self, conn: Any, item: dict[str, Any]) -> None:
        combined_text = " ".join(
            part
            for part in [
                item["name"],
                item["brand"],
                item["category"],
                item["spec"],
                item["description"],
                item["selling_points"],
            ]
            if part
        )
        await conn.execute(
            """
            INSERT INTO product_knowledge (
                spu_id, sku_id, name, category, brand, spec, description,
                image_text, combined_text, image_urls, price, status, updated_at
            ) VALUES (
                $1, $2, $3, NULLIF($4, ''), NULLIF($5, ''), NULLIF($6, ''), NULLIF($7, ''),
                '', $8, '{}', $9, $10, NOW()
            )
            ON CONFLICT (spu_id, sku_id) DO UPDATE SET
                name = EXCLUDED.name,
                category = COALESCE(NULLIF(EXCLUDED.category, ''), product_knowledge.category),
                brand = COALESCE(NULLIF(EXCLUDED.brand, ''), product_knowledge.brand),
                spec = COALESCE(NULLIF(EXCLUDED.spec, ''), product_knowledge.spec),
                description = COALESCE(NULLIF(EXCLUDED.description, ''), product_knowledge.description),
                combined_text = EXCLUDED.combined_text,
                price = COALESCE(EXCLUDED.price, product_knowledge.price),
                status = EXCLUDED.status,
                updated_at = NOW()
            """,
            item["spu_id"],
            item["sku_id"],
            item["name"],
            item["category"],
            item["brand"],
            item["spec"],
            item["description"],
            combined_text or item["name"],
            item["retail_price"],
            item["status"],
        )

    def _aggregate_inventory_records(self, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        aggregated: dict[str, dict[str, Any]] = {}
        for item in items:
            sku_id = item.get("sku_id")
            if not sku_id:
                continue
            current = aggregated.get(sku_id)
            if current is None:
                current = {
                    "sku_id": sku_id,
                    "product_name": item.get("product_name", ""),
                    "barcode": item.get("barcode", ""),
                    "category": item.get("category", ""),
                    "spec": item.get("spec", ""),
                    "cost_price": item.get("cost_price"),
                    "stock": 0,
                    "store_ids": set(),
                    "store_names": set(),
                    "warehouses": set(),
                }
                aggregated[sku_id] = current
            current["stock"] += int(item.get("stock") or 0)
            if not current["product_name"] and item.get("product_name"):
                current["product_name"] = item["product_name"]
            if not current["barcode"] and item.get("barcode"):
                current["barcode"] = item["barcode"]
            if not current["category"] and item.get("category"):
                current["category"] = item["category"]
            if not current["spec"] and item.get("spec"):
                current["spec"] = item["spec"]
            if current["cost_price"] is None and item.get("cost_price") is not None:
                current["cost_price"] = item["cost_price"]
            if item.get("store_id"):
                current["store_ids"].add(item["store_id"])
            if item.get("store_name"):
                current["store_names"].add(item["store_name"])
            if item.get("warehouse"):
                current["warehouses"].add(item["warehouse"])

        normalized = []
        for current in aggregated.values():
            store_ids = sorted(current["store_ids"])
            store_names = sorted(current["store_names"])
            warehouses = sorted(current["warehouses"])
            normalized.append(
                {
                    **current,
                    "store_ids": store_ids,
                    "store_names": store_names,
                    "warehouses": warehouses,
                    "store_id": store_ids[0] if len(store_ids) == 1 else None,
                }
            )
        return normalized

    async def _run_post_import_etl(self, import_type: str) -> None:
        """导入完成后自动跑派生 ETL，确保下游数据表同步更新。"""
        logger.info("🔄 运行 post-import ETL (type=%s)...", import_type)

        # 1. 销售历史聚合（订单导入后必跑）
        if import_type in ("orders", "products"):
            try:
                from src.sync.etl_sales_aggregation import run_sales_aggregation_etl
                result = await run_sales_aggregation_etl(self._pool)
                logger.info("✅ 销售历史聚合: %s", result)
            except ImportError:
                logger.debug("etl_sales_aggregation 不可用，跳过")
            except Exception as e:
                logger.warning("⚠️ 销售历史聚合失败: %s", e)

        # 2. 类目映射（商品导入后必跑）
        if import_type == "products":
            try:
                from src.sync.etl_category_mapping import run_category_mapping_etl
                result = await run_category_mapping_etl(self._pool, None)
                logger.info("✅ 类目映射: %s", result)
            except Exception as e:
                logger.warning("⚠️ 类目映射失败: %s", e)

        # 3. 商品关联挖掘（订单导入后必跑）
        if import_type == "orders":
            try:
                from src.sync.etl_product_associations import run_product_associations_etl
                result = await run_product_associations_etl(self._pool)
                logger.info("✅ 商品关联: %s", result)
            except Exception as e:
                logger.warning("⚠️ 商品关联失败: %s", e)

        logger.info("✅ post-import ETL 完成")

    async def _record_run(
        self,
        run_id: str,
        import_type: str,
        filename: str,
        detected_sheets: list[str],
        total_rows: int,
        imported_rows: int,
        skipped_rows: int,
        quality_report: dict[str, Any],
        import_summary: dict[str, Any],
    ) -> None:
        await self.pool.execute(
            """
            INSERT INTO manual_import_runs (
                run_id, import_type, filename, status, detected_sheets, total_rows,
                imported_rows, skipped_rows, quality_score, quality_report, import_summary,
                created_at, updated_at
            ) VALUES (
                $1, $2, $3, 'completed', $4::text[], $5, $6, $7, $8, $9::jsonb, $10::jsonb,
                NOW(), NOW()
            )
            """,
            run_id,
            import_type,
            filename,
            detected_sheets,
            total_rows,
            imported_rows,
            skipped_rows,
            quality_report.get("score", 0),
            json.dumps(quality_report, ensure_ascii=False, default=str),
            json.dumps(import_summary, ensure_ascii=False, default=str),
        )

    def _build_quality_report(
        self,
        total_rows: int,
        issues: list[QualityIssue],
        stats: dict[str, Any],
        suggestions: list[str],
    ) -> dict[str, Any]:
        severity_weight = {"critical": 5, "warning": 3, "info": 1}
        weighted_issues = sum(severity_weight.get(issue.severity, 1) * issue.count for issue in issues)
        return {
            "score": _quality_score(total_rows, weighted_issues),
            "stats": stats,
            "issues": [issue.to_dict() for issue in issues],
            "suggestions": suggestions,
        }

    def _normalize_platform(self, raw: str) -> str:
        mapping = {
            "美团闪购": "meituan",
            "淘宝闪购": "taobao_flash",
            "京东到家": "jddj",
            "饿了么": "eleme",
        }
        return mapping.get(raw, raw.lower() if raw else "manual")

    def _extract_phone_suffix(self, raw: str) -> str:
        digits = "".join(ch for ch in raw if ch.isdigit())
        return digits[-4:] if digits else ""
